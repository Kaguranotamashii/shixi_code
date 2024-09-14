import re

import requests
import json
import openpyxl

# 定义请求的URL
url = "http://192.168.2.5:8000/similarity_search"

# 定义请求的headers
headers = {
    'Content-Type': 'application/json'
}
I=0
# 读取demo.txt文件
with open('demo5', 'r', encoding='utf-8') as file:
    queries = file.readlines()

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Search Results"

# 设置表头
ws.cell(row=1, column=1, value="Query")
ws.cell(row=1, column=2, value="Source")
ws.cell(row=1, column=3, value="meta_data")
# 遍历每一行查询
row_index = 2  # 从第二行开始写入数据
for query in queries:
    query = query.strip()  # 去除首尾空白字符
    payload = {
        "query": query,
        "top_k": 1000,
        "search_score": 0.95
    }
    json_payload = json.dumps(payload)

    # 发送POST请求
    response = requests.post(url, data=json_payload, headers=headers)
    source1= ""
    # 这里定义给集合不能有重复的set
    source_set = set()

    I+=1




    meta_data1=""
    # 检查响应状态码
    if response.status_code == 200:
        response_data = response.json()
        if isinstance(response_data, list):
            for item in response_data:
                if 'source' in item:
                    source = item['source']
                    # 去除路径前缀
                    if source.startswith("D:\\ceshi\\AIPCBackend\\files\\"):
                        source = source[len("D:\\ceshi\\AIPCBackend\\files\\"):]
                        if source not in source_set:
                            source_set.add(source)
                            source1 += source + " \n"
                            print( source1)

                    # 将查询和提取的source写入Excel
                if 'meta_data' in item:
                    meta_data = item['meta_data']
                    # 替换非法字符
                    meta_data = re.sub(r'\s+|[^\w\s]', '', meta_data)
                    # 去除路径前缀
                    meta_data1+=meta_data+" \n"


        else:
            print(f"Unexpected response format for query '{query}': {response_data}")
    else:
        print(f"Failed to get response for query '{query}', status code: {response.status_code}")
    ws.cell(row=row_index, column=1, value=query)
    ws.cell(row=row_index, column=2, value=source1)
    ws.cell(row=row_index, column=3, value=meta_data1)
    row_index += 1
    print(I)

# 保存Excel文件
wb.save('search_results4.xlsx')
print("Results saved to search_results.xlsx")