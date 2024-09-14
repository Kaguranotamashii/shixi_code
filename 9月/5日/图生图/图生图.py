import os
import requests
import base64
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# 设置API URL
url = "http://10.27.0.2:60000/image_to_image"

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "Image Results"

# 调整列宽
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30

# 图片文件夹
img_dir = "./img"

# 设置每个图片嵌入的大小（可调整）
w=1024
h=1024
prompt="男人的头发变为红色"
num=2
img_size = (w, h)

# 遍历img文件夹中的图片
for img_file in os.listdir(img_dir):
    if img_file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
        # 读取图片
        img_path = os.path.join(img_dir, img_file)

        # 设置POST请求数据
        files = {
            'file': (img_file, open(img_path, 'rb'), 'image/jpeg')  # 图片类型可以根据实际格式调整
        }
        data = {
            'strength': '0.5',
            'guidance_scale': '5',
            'prompt': prompt,
            'width': w,
            'height': h,
            'num_inference_steps': '50',
            'seed': '42',
            'num_images_per_prompt': num
        }

        # 发送请求
        response = requests.post(url, headers={'accept': 'application/json'}, files=files, data=data)
        print(response.text)


        if response.status_code == 200:

            json_response = response.json()

            # 从响应中获取生成的base64图片
            generated_images = []
            for img_data_base64 in json_response.get('images', []):
                img_data = base64.b64decode(img_data_base64)
                img = Image.open(BytesIO(img_data))
                generated_images.append(img)

            # 调整原图大小并嵌入Excel单元格
            img_original = Image.open(img_path)
            img_original.thumbnail(img_size)  # 调整大小
            img_original_save_path = f'原图_{prompt}_{img_file}'
            img_original.save(img_original_save_path)

            xl_img_original = XLImage(img_original_save_path)
            ws.add_image(xl_img_original, f'A{ws.max_row + 1}')  # 插入到当前行的A列

            # 插入生成的图片
            for i, gen_img in enumerate(generated_images):
                # 将生成的图片调整大小并保存
                gen_img.thumbnail(img_size)
                img_save_path = f'图生图后的_{img_file.split(".")[0]}_{i}_{prompt}.png'
                gen_img.save(img_save_path)

                xl_img_generated = XLImage(img_save_path)
                ws.add_image(xl_img_generated, f'B{ws.max_row}')  # 插入到当前行的B列

        else:
            print(f"Error processing {img_file}: {response.status_code}")

# 保存Excel文件
wb.save("Image_Results.xlsx")
